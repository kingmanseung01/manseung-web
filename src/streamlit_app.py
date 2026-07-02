import os
# 프로그램이 켜질 때 openpyxl이 없으면 지가 알아서 강제로 설치하는 치트키 코드
os.system("pip install openpyxl")
import streamlit as st  
import os
import io
from datetime import datetime
from openpyxl import load_workbook

# --- [일련번호 관리 및 날짜 초기화 로직] ---
today_str = datetime.today().strftime('%y%m%d')

if "last_date" not in st.session_state:
    st.session_state.last_date = today_str

if "document_counter" not in st.session_state:
    st.session_state.document_counter = 1

if st.session_state.last_date != today_str:
    st.session_state.last_date = today_str
    st.session_state.document_counter = 1


# --- [데이터 기본 세팅 (대표사 & 성상 목록)] ---
if "company_list" not in st.session_state:
    st.session_state.company_list = [
        "주식회사 파란에코(운반)",
        "주식회사 하나씨엔알(운반)",
        "주식회사 지엘산업개발(운반)",
        "주식회사 장형기업(처리)",
        "주식회사 한성기업(처리)",
        "주식회사 장형지피(처리)",
        "한밭미래자원 주식회사(처리)",
        "주식회사 장형지아이(처리)"
    ]

if "waste_type_list" not in st.session_state:
    st.session_state.waste_type_list = [
        "폐콘크리트",
        "폐아스팔트콘크리트",
        "혼합건설폐기물",
        "폐합성수지",
        "폐목재",
        "폐보드 및 폐판넬",
        "그밖의 폐기물"
    ]

if "step" not in st.session_state:
    st.session_state.step = 1

if "selected_company" not in st.session_state:
    st.session_state.selected_company = st.session_state.company_list[0]


# --- [사이드바 메뉴] ---
menu = st.sidebar.selectbox(
    "원하는 작업을 선택하세요",
    ["홈", "착수계 관리", "준공계 관리", "청구서 관리"]
)

if "current_menu" not in st.session_state:
    st.session_state.current_menu = menu
elif st.session_state.current_menu != menu:
    st.session_state.current_menu = menu
    st.session_state.step = 1


if menu == "홈":
    st.title("💻 업무 자동화 웹 시스템")
    st.write("인터넷만 연결되어 있으면 어디서든 사용할 수 있는 업무 자동화 공간입니다.")

elif menu == "착수계 관리":
    st.title("📝 착수계 작성 및 관리")
    
    # --- [1단계 화면: 대표사만 선택] ---
    if st.session_state.step == 1:
        st.subheader("1단계: 대표사 선택 및 관리")
        
        selected = st.selectbox(
            "대표사를 골라주세요",
            st.session_state.company_list,
            index=st.session_state.company_list.index(st.session_state.selected_company) if st.session_state.selected_company in st.session_state.company_list else 0
        )
        st.session_state.selected_company = selected
        
        st.markdown("---")
        
        with st.expander("🛠️ 대표사 목록 관리 (추가 / 삭제)"):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**[ 대표사 추가 ]**")
                new_company = st.text_input("추가할 회사명 입력", key="add_input", placeholder="예: 주식회사 신규에코(운반)")
                if st.button("목록에 추가하기"):
                    if new_company and new_company not in st.session_state.company_list:
                        st.session_state.company_list.append(new_company)
                        st.success(f"'{new_company}' 추가 완료!")
                        st.rerun()
            with col2:
                st.markdown("**[ 대표사 삭제 ]**")
                delete_target = st.selectbox("삭제할 회사 선택", st.session_state.company_list, key="del_select")
                if st.button("목록에서 삭제하기"):
                    if len(st.session_state.company_list) > 1:
                        st.session_state.company_list.remove(delete_target)
                        st.success(f"'{delete_target}' 삭제 완료!")
                        st.rerun()
                        
        st.markdown("---")
        
        if st.button("대표사 선택 완료 ➡️ 다음 단계로"):
            st.session_state.step = 2
            st.rerun()

    # --- [2단계 화면: 상세 정보 입력] ---
    elif st.session_state.step == 2:
        st.subheader("2단계: 착수계 상세 정보 및 성상 입력")
        st.info(f"🏢 선택된 대표사: **{st.session_state.selected_company}**")
        
        current_serial = f"{today_str}-{st.session_state.document_counter:02d}"
        st.warning(f"💡 이번 문서의 착수 번호는 **{current_serial}**로 자동 생성됩니다.")

        if st.button("⬅️ 대표사 다시 고르기 (이전으로)"):
            st.session_state.step = 1
            st.rerun()
            
        st.markdown("---")
        
        client_name = st.text_input("1. 배출자", placeholder="예) 인천광역시 제물포구청")
        
        st.markdown("### 🌿 2. 성상 관리 및 선택")
        with st.expander("🛠️ 성상 목록 자체를 추가/삭제하려면 여기를 클릭하세요"):
            col3, col4 = st.columns(2)
            with col3:
                new_waste = st.text_input("새로운 성상 추가", placeholder="예: 폐목재(임목)")
                if st.button("성상 목록에 추가"):
                    if new_waste and new_waste not in st.session_state.waste_type_list:
                        st.session_state.waste_type_list.append(new_waste)
                        st.success(f"'{new_waste}' 목록 추가 완료!")
                        st.rerun()
            with col4:
                delete_waste_target = st.selectbox("기존 성상 삭제", st.session_state.waste_type_list)
                if st.button("성상 목록에서 제거"):
                    if len(st.session_state.waste_type_list) > 1:
                        st.session_state.waste_type_list.remove(delete_waste_target)
                        st.success(f"'{delete_waste_target}' 목록 제거 완료!")
                        st.rerun()

        selected_wastes = st.multiselect(
            "이 문서에 포함할 성상을 모두 골라주세요 (여러 개 선택 가능)",
            st.session_state.waste_type_list,
            default=[st.session_state.waste_type_list[0]]
        )
        
        waste_tons = {}
        total_ton = 0.0
        
        if selected_wastes:
            st.markdown("##### ⚖️ 고른 성상별 계약톤수 입력")
            for waste in selected_wastes:
                ton_val = st.number_input(
                    f"└─ {waste} 톤수입력", 
                    min_value=0.0, 
                    value=0.0, 
                    step=0.01, 
                    format="%.2f",
                    key=f"ton_{waste}"
                )
                waste_tons[waste] = ton_val
                total_ton += ton_val
            
            detail_strings = [f"{w}({t:.2f}톤)" for w, t in waste_tons.items()]
            waste_result_string = ", ".join(detail_strings)
            st.info(f"📊 **총 톤수 합계:** {total_ton:.2f} 톤")
        else:
            waste_result_string = ""
            st.error("⚠️ 성상을 최소 한 개 이상 선택해 주세요!")
            
        st.markdown("---")
        
        project_name = st.text_input("3. 용역명", placeholder="예: 대호만 폐기물 처리 용역")
        contract_price = st.number_input("4. 계약금액 (원)", min_value=0, value=0, step=10000)
        
        contract_date = st.date_input("5. 계약일")
        start_date = st.date_input("6. 착수일")
        end_date = st.date_input("7. 준공예정일")
        actual_end_date = st.date_input("8. 실제준공일")
        
        st.markdown("---")
        
        # 입력 검증
        input_ready = False
        if client_name.strip() and project_name.strip() and contract_price > 0 and total_ton > 0.0 and selected_wastes:
            input_ready = True

        if not input_ready:
            st.warning("⚠️ 모든 빈칸을 채우고 계약금액 및 톤수를 입력하면 다운로드 버튼이 활성화됩니다.")
        else:
            current_company = st.session_state.selected_company
            template_path = ""

            if "파란에코" in current_company:
                template_path = "template_paran.xlsx"
            elif "한성기업" in current_company:
                template_path = "template_hansung.xlsx"
            elif "한밭미래자원" in current_company:
                template_path = "template_hanbat.xlsx"
            else:
                template_path = "template_paran.xlsx" 
            
            output_filename = f"착수계_{client_name}_{current_serial}.xlsx"

            # 도커 기반 허깅페이스 서버에서는 파일이 루트 폴더(/)에 위치할 수 있으므로 경로 체크
            if not os.path.exists(template_path):
                # 루트 폴더에 없고 src 폴더 안에 있을 가능성 대비
                alt_path = os.path.join("src", template_path)
                if os.path.exists(alt_path):
                    template_path = alt_path

            if not os.path.exists(template_path):
                st.error(f"⚠️ 서버 폴더 내에 '{template_path}' 파일이 없습니다. Files 탭에서 양식 엑셀 파일들을 꼭 업로드해 주세요.")
            else:
                try:
                    # openpyxl을 이용한 메모리 기반 데이터 주입
                    wb = load_workbook(template_path, data_only=False)
                    
                    if "template_paran.xlsx" in template_path:
                        ws_first = wb.worksheets[0]
                        ws_first['D6'].value = current_serial
                        
                        if '기초정보' in wb.sheetnames:
                            ws_info = wb['기초정보']
                        else:
                            ws_info = wb.worksheets[8]
                        
                        ws_info['B2'].value = project_name
                        ws_info['B4'].value = contract_price
                        ws_info['B5'].value = str(contract_date)
                        ws_info['B6'].value = str(start_date)
                        ws_info['B7'].value = str(end_date)
                        ws_info['B8'].value = str(actual_end_date)
                        ws_info['B9'].value = client_name
                        ws_info['B10'].value = total_ton
                        ws_info['B11'].value = waste_result_string
                        
                        ws_third = wb.worksheets[2]
                        start_row = 8
                        for i, (waste_name, ton_val) in enumerate(waste_tons.items()):
                            current_row = start_row + i
                            ws_third[f'D{current_row}'].value = waste_name
                            ws_third[f'G{current_row}'].value = ton_val
                    else:
                        ws = wb.worksheets[0]
                        ws['B3'].value = client_name
                        ws['B4'].value = project_name

                    # 엑셀 파일을 물리적으로 저장하지 않고 메모리(Stream)에 담기
                    excel_stream = io.BytesIO()
                    wb.save(excel_stream)
                    excel_stream.seek(0)
                    
                    st.success("✨ 엑셀 파일이 성공적으로 준비되었습니다! 아래 버튼을 눌러 다운로드하세요.")
                    
                    def increment_counter():
                        st.session_state.document_counter += 1

                    st.download_button(
                        label="📥 엑셀 파일 다운로드 받기",
                        data=excel_stream,
                        file_name=output_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        on_click=increment_counter
                    )
                    
                except Exception as e:
                    st.error(f"엑셀 처리 중 에러 발생: {e}")

elif menu == "준공계 관리":
    st.title("🗂️ 준공계 작성 및 관리")
    st.write("준공계 페이지")

elif menu == "청구서 관리":
    st.title("💰 청구서 작성 및 관리")
    st.write("청구서 페이지")
    
